import openpyxl
import canvas_integration

def create_workbook(name, users, canvas):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row = 1, column = 1).value = "Student"
    sheet.cell(row = 1, column = 2).value = "Til stede"

    for index, user in enumerate(users, 2):
        sheet.cell(row = index, column = 1).value = canvas.users[user].short_name

    wb.save(filename = name + ".xlsx")


if __name__ == "__main__":
    c = canvas_integration.CanvasIntegration("https://uit.instructure.com", canvas_integration.get_token("canvas-token"), "18522")
    sections = c.get_section_overview()
    for section in sections:
        create_workbook(section, sections[section], c)
